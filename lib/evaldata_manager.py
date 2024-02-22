import os
import json
import openpyxl
import pandas as pd
from pandas.api.types import is_numeric_dtype


def get_model_runspec_map(root):
    """
    Does Directory traversal from root
    :param root:
    :return: Returns Model -> Runspecs Map and Runspec -> Models Map
    """
    models = {}
    runspecs = {}
    for model in os.listdir(root):
        if not os.path.isdir(os.path.join(root, model)):
            continue
        if model not in models:
            models[model] = set()
        for datafile in os.listdir(os.path.join(root, model)):
            parts = datafile.split('.')
            nametag = '.'.join(parts[:-1])
            ext = parts[-1]
            if ext != 'json':
                continue
            if datafile.startswith("~"):
                continue
            nametag = nametag.replace(model + '_', '')
            nametag = "runspec" + nametag.split('runspec')[1]
            if "_overall_metrics_" in nametag:
                runspec, when = nametag.split('_overall_metrics_')
                if runspec not in runspecs:
                    runspecs[runspec] = set()
                runspecs[runspec].add(model)
                models[model].add(runspec)
    for model in models:
        models[model] = sorted(list(models[model]), key=str.casefold)
    for runspec in runspecs:
        runspecs[runspec] = sorted(list(runspecs[runspec]), key=str.casefold)
    return models, runspecs


def get_best_matched_common_fields(df_columns, commom_fields_map):
    cf_fields = ''
    cf_count = {}
    for key in commom_fields_map:
        cf_fields = commom_fields_map.split(',')
        # print(key, common_fields)
        for cf in cf_fields:
            # print("Checking", cf, "in", list(df_columns))
            if cf in df_columns:
                if commom_fields_map[key] not in cf_count:
                    cf_count[commom_fields_map[key]] = 0
                cf_count[commom_fields_map[key]] += 1
    cf_max = 0
    for ckey in cf_count:
        if cf_count[ckey] > cf_max:
            cf_fields = ckey.split(',')
            cf_max = cf_count[ckey]
    return cf_fields


def overall_data_postprocessing(runspec, attrs):
    if runspec in ["runspec_summary_case", "runspec_summary_cnndm"]:
        delkeys = []
        newdata = {}
        for key in attrs:
            if isinstance(attrs[key], dict):
                delkeys.append(key)
                for label in attrs[key]:
                    newdata[key + '_' + label] = attrs[key][label]
        attrs = newdata
    if runspec == "runspec_gender_biasness_winobias":
        for key in attrs:
            if isinstance(attrs[key], dict):
                for x in attrs[key]:
                    attrs[key] = attrs[key][x]
    return attrs


def add_attrs(attrs, model, runspec, when, data):
    attrs["model"] = model
    attrs["run_spec"] = runspec
    attrs["run_time"] = when
    if runspec not in data:
        data[runspec] = []
    append_data = True
    for mdata in data[runspec]:
        if mdata["model"] == model:
            if mdata["run_time"] >= when:
                append_data = False
            else:
                data[runspec].remove(mdata)
    if append_data:
        data[runspec].append(attrs)
    return data


def get_run_spec_details(datafile, model, root):
    # print("Datafile:", datafile)
    parts = datafile.split('.')
    nametag = '.'.join(parts[:-1])
    ext = parts[-1]
    # print("NameTag 1:", nametag)
    nametag = nametag.replace(model + '_', '')
    # print("NameTag 2:", nametag)
    nameleft = "unknown"
    attrs = {}
    if ext == "json":
        if "_overall_metrics_" in nametag:
            nameleft, when = nametag.split('_overall_metrics_')
            type = "overall"
            with open(os.path.join(root, model, datafile)) as fh:
                attrs = json.load(fh)
        elif "_response_" in nametag:
            nameleft, when = nametag.split('_response_')
            nameleft = '_'.join(nameleft.split('_')[:-1])
            type = "response"
            with open(os.path.join(root, model, datafile)) as fh:
                attrs = json.load(fh)
    elif ext == "xlsx":
        if "_record_level_metrics_" in nametag:
            nameleft, when = nametag.split('_record_level_metrics_')
        elif "_response_" in nametag:
            nameleft, when = nametag.split('_response_')
        else:
            print("Excel NameTag", nametag, "processing not implemented")

        # nameleft = '_'.join(nameleft.split('_')[:-1])
        type = "metrics"
        try:
            wb = openpyxl.load_workbook(os.path.join(root, model, datafile))
        except Exception as e:
            print("Excel %s/%s failed: %s" % (model, datafile, str(e)))
            return nameleft, "unknown", "unknown", {}
        ws = wb.active
        cols = [c.value for c in ws[1]]
        records = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            record = {}
            for i, c in enumerate(cols):
                record[c] = row[i]
            records.append(record)
        # attrs["records"] = records
        attrs = records
    else:
        print(datafile, "did not fall in any bucket")
        type = "unknown"
        when = "unknown"
    if len(nameleft.split("runspec")) > 2:
        print("Logic Failed for", nameleft)
    return "runspec" + nameleft.split("runspec")[1], when, type, attrs


def load_data_from_files(root, models, runspecs):
    outcome_data = {}
    record_data = {}
    for model in models:
        if not os.path.isdir(os.path.join(root, model)):
            continue
        for datafile in os.listdir(os.path.join(root, model)):
            ext = datafile.split('.')[-1]
            if ext not in ['json', 'xlsx']:
                continue
            if datafile.startswith("~"):
                continue
            runspec, when, type, attrs = get_run_spec_details(datafile, model, root)
            if runspec not in runspecs:
                continue
            if type == "overall":
                attrs = overall_data_postprocessing(runspec, attrs)
                outcome_data = add_attrs(attrs, model, runspec, when, outcome_data)
            else:
                if runspec not in record_data:
                    record_data[runspec] = {}
                record_data[runspec][model] = attrs
    return outcome_data, record_data


def convert_runspec_to_df(dict_data):
    for runspec in dict_data:
        dict_data[runspec] = pd.DataFrame(dict_data[runspec])
        dict_data[runspec] = dict_data[runspec].apply(pd.to_numeric, errors='ignore')


def get_runspec_record_common_fields(runspec, df, commom_fields_map):
    if runspec in commom_fields_map:
        print("Set On field", commom_fields_map[runspec])
        common_fields = commom_fields_map[runspec].split(',')
    else:
        # This is the most common on fields so lets try it
        common_fields = get_best_matched_common_fields(list(df.columns), commom_fields_map)
        print("Matched On Fields", common_fields, "\nColumns", list(df.columns))
    print("Common Fields:", common_fields)
    return common_fields


def rename_df_columns(common_fields, df, model_tag):
    new_names = {}
    for cf in df.columns:
        if cf not in common_fields:
            new_names[cf] = cf + '_' + model_tag
    df.rename(columns=new_names, inplace=True)


def denormalize_records_data(record_data):
    denormalized_recorddata = {}
    for runspec in record_data:
        records = []
        for model in record_data[runspec]:
            data = record_data[runspec][model]
            for record in data:
                record["runspec"] = runspec
                record["model"] = model
                records.append(record)
        if runspec not in denormalized_recorddata:
            denormalized_recorddata[runspec] = {}
        denormalized_recorddata[runspec] = records
    return denormalized_recorddata


def get_data(root, models, runspecs, commom_fields_map):
    print("Get Data: Models %s RunSpecs %s" % (len(models), len(runspecs)))
    outcome_data, record_data = load_data_from_files(root, models, runspecs)
    convert_runspec_to_df(outcome_data)

    for runspec in record_data:
        applicable_models = [model for model in models if model in record_data[runspec]]
        # No model ran this runspec
        if len(applicable_models) == 0:
            continue
        base_model = applicable_models[0]
        df = pd.DataFrame(record_data[runspec][base_model])
        # Only 1 model ran this runspec
        if len(applicable_models) == 1:
            record_data[runspec] = df
            continue
        # Multiple models ran this runspec
        common_fields = get_runspec_record_common_fields(runspec, df)
        common_headers = []
        for col in record_data[runspec][base_model][0].keys():
            if col in common_fields:
                common_headers.append(col)
        print(base_model, "Common Headers:", common_headers, "matching", record_data[runspec][base_model][0].keys())
        if len(common_headers) == 0:
            # get_data_using_concat(runspec, record_data, models)
            print("No common column found between " + runspec + " runs")
            print("Expecting atleast one of the column[%s] to exist" % ','.join(common_fields))
            print("Among the records headers " + ' '.join(list(record_data[runspec][base_model][0].keys())))
            print("Ask admin to add common fields for " + runspec + " in ini config file")
        else:
            rename_df_columns(common_fields, df, base_model)
            for model in applicable_models[1:]:
                other_df = pd.DataFrame(record_data[runspec][model])
                rename_df_columns(common_fields, other_df, model)
                try:
                    df = df.merge(other_df, on=common_headers)
                    # print("Merge Successful")
                except Exception as e:
                    print(model, runspec, "Df Columns:", df.columns)
                    print(model, runspec, "Other DF Columns:", other_df.columns)
                    print(model, runspec, "Merge Failed:", str(e))
            # Rearrange the columns
            cols = common_headers + sorted(list(set(df.columns) - set(common_headers)))
            df = df[cols]
            record_data[runspec] = df
    for runspec in record_data:
        record_data[runspec] = record_data[runspec].apply(pd.to_numeric, errors='ignore')

    return outcome_data, record_data